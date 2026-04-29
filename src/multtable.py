#!/usr/bin/env python3

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

fill = PatternFill(start_color="AAAAFF", end_color="AAAAFF", fill_type="solid")

wb = openpyxl.Workbook()

sheet = wb.active
sheet.title = "Multiplication Table"

for v in range(1, 11):
  for cell in [sheet.cell(row=1, column=1+v), sheet.cell(row=1+v, column=1)]:
    cell.value = v
    cell.font = Font(bold=True)
  sheet.cell(row=1+v, column=1+v).fill = fill

for y in range(1, 11):
  for x in range(1, 11):
    cell = sheet.cell(row=1+y, column=1+x)
    cell.value = '=%s%s*%s%s' % (get_column_letter(x+1), 1, get_column_letter(1), y+1)

wb.save("multtable.xlsx")
